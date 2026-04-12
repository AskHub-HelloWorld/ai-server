"""AskHub AI Server WBS 엑셀 생성 스크립트."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "WBS"

# ── 색상/스타일 정의 ──
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
PHASE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
PHASE_FONT = Font(name="맑은 고딕", bold=True, size=10)
DONE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
TODO_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
GANTT_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
GANTT_DONE_FILL = PatternFill(start_color="A9D18E", end_color="A9D18E", fill_type="solid")
CELL_FONT = Font(name="맑은 고딕", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── 컬럼 정의 ──
COLUMNS = ["ID", "작업명", "기간(일)", "선행작업", "산출물", "상태", "W1", "W2", "W3", "W4", "W5"]
COL_WIDTHS = [8, 45, 8, 12, 45, 8, 6, 6, 6, 6, 6]

# ── 헤더 작성 ──
for col_idx, (name, width) in enumerate(zip(COLUMNS, COL_WIDTHS), 1):
    cell = ws.cell(row=1, column=col_idx, value=name)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = THIN_BORDER
    cell.alignment = CENTER
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# ── 데이터 정의 ──
# (id, name, duration, dependency, deliverable, status, gantt_weeks)
# gantt_weeks: "d1" = 완료(초록), 숫자 = 미완료 주차(파랑)
TASKS = [
    # === P0 ===
    ("phase", "P0. 개발환경 설정", None, None, None, "완료", []),
    ("P0-1", "프로젝트 구조 생성 (Python/FastAPI)", "-", "-", "pyproject.toml, src/ 구조", "완료", ["d1"]),
    ("P0-2", "서버 상태 확인 API 개발", "-", "P0-1", "api/routes/health.py", "완료", ["d1"]),
    ("P0-3", "Docker 환경 구성", "-", "P0-1", "Dockerfile, compose.yaml, .env.example", "완료", ["d1"]),
    ("P0-4", "기본 테스트 작성", "-", "P0-2", "tests/test_health.py", "완료", ["d1"]),

    # === P1 ===
    ("phase", "P1. 기본 AI 채팅", None, None, None, "완료", []),
    ("P1-1", "LLM 서비스 구현 (Bedrock Nova Lite)", "-", "P0-3", "services/llm.py", "완료", ["d1"]),
    ("P1-2", "실시간 응답 스트리밍 구현 (SSE)", "-", "P1-1", "api/routes/chat.py", "완료", ["d1"]),
    ("P1-3", "RAG 소스/작업 API 기본 구조 생성", "-", "P1-1", "api/routes/sources.py, ingestion_jobs.py", "완료", ["d1"]),
    ("P1-4", "이전 채팅 API 제거", "-", "P1-2", "POST /v1/chat, /v1/chat/stream 삭제", "완료", ["d1"]),

    # === P2 ===
    ("phase", "P2. DB 연결 + 채팅 히스토리", None, None, None, "완료", []),
    ("P2-1", "PostgreSQL 연결 설정", "-", "P1-1", "core/database.py, core/config.py", "완료", ["d1"]),
    ("P2-2", "ai 전용 스키마 분리 + DB 버전 관리 설정", "-", "P2-1", "alembic/, alembic.ini", "완료", ["d1"]),
    ("P2-3", "채팅 세션/메시지 DB 모델 생성", "-", "P2-2", "models/chat.py", "완료", ["d1"]),
    ("P2-4", "채팅 테이블 생성 (DB 버전 관리)", "-", "P2-3", "alembic/versions/202604070001_...", "완료", ["d1"]),
    ("P2-5", "세션 기반 채팅 API 개발 (5개)", "-", "P2-4", "api/routes/chat.py, schemas/chat.py", "완료", ["d1"]),
    ("P2-6", "대화 이력 DB 직접 관리 기능", "-", "P2-5", "api/routes/chat.py 내 조회/저장 로직", "완료", ["d1"]),
    ("P2-7", "채팅 기능 통합 테스트", "-", "P2-6", "tests/test_chat_sessions.py", "완료", ["d1"]),

    # === P2-후속 ===
    ("phase", "P2-후속. 파일 관리", None, None, None, "완료", []),
    ("P2F-1", "업로드 파일 DB 모델 생성", "-", "P2-4", "models/file.py", "완료", ["d1"]),
    ("P2F-2", "파일 테이블 생성 (DB 버전 관리)", "-", "P2F-1", "alembic/versions/202604080001_...", "완료", ["d1"]),
    ("P2F-3", "파일 업로드/조회 API 개발", "-", "P2F-2", "api/routes/files.py, schemas/file.py", "완료", ["d1"]),
    ("P2F-4", "첨부 파일 내용을 AI 답변에 반영", "-", "P2F-3", "api/routes/chat.py 수정", "완료", ["d1"]),
    ("P2F-5", "EC2 환경 Bedrock 연결 설정", "-", "-", "core/config.py 수정", "완료", ["d1"]),

    # === P3 ===
    ("phase", "P3. RAG 검색 (키워드 + 벡터 혼합)", None, None, None, "미착수", []),
    ("P3-1", "DocumentChunk DB 모델 추가\n(RagSource/IngestionJob metadata 모델은 완료)", "1", "-", "models/document.py", "미착수", [1]),
    ("P3-2", "document_chunks 테이블 생성 (DB 버전 관리)\n(벡터 컬럼 + 키워드 검색 컬럼 + 검색 인덱스)", "1", "P3-1", "alembic/versions/..._create_document_chunks.py", "미착수", [1]),
    ("P3-3", "임베딩 관련 설정 추가\n(모델 ID, 벡터 차원 수)", "0.5", "-", "core/config.py 수정", "미착수", [1]),
    ("P3-4", "임베딩 서비스 구현\n(Bedrock Titan Embed V2로 텍스트를 벡터로 변환)", "1.5", "P3-3", "services/embedding.py", "미착수", [1, 2]),
    ("P3-5", "혼합 검색 서비스 구현\n(키워드 검색 + 벡터 유사도 검색 결합, 팀별 필터)", "2", "P3-2, P3-4", "services/retriever.py", "미착수", [2]),
    ("P3-6", "답변 가능 여부 판단 서비스 구현\n(검색 결과가 부족하면 질문글 초안 작성)", "1", "P3-5", "services/answer_policy.py", "미착수", [2, 3]),
    ("P3-7", "출처 표기 서비스 구현\n(검색된 문서/코드의 출처 정보를 응답에 포함)", "1", "P3-5", "services/citation_builder.py", "미착수", [2, 3]),
    ("P3-8", "ChatService 기반 RAG 통합\n(RAG 검색 결과를 AI 답변에 반영)", "2", "P3-5, P3-6,\nP3-7", "services/chat_service.py,\napi/routes/chat.py 수정", "미착수", [3]),
    ("P3-9", "RAG 통합 테스트\n(검색 -> 출처 포함 답변 -> 답변 불가 시 질문글 초안 검증)", "1", "P3-8", "테스트 시나리오 검증 완료", "미착수", [3]),

    # === P4 ===
    ("phase", "P4. 문서 수집 및 저장", None, None, None, "미착수", []),
    ("P4-1", "소스/작업 API를 worker 처리 흐름과 연결", "1.5", "P3-2", "api/routes/sources.py,\ningestion_jobs.py 수정", "미착수", [3, 4]),
    ("P4-2", "백그라운드 작업 처리기 구현\n(DB에서 대기 중인 작업을 가져와 순차 처리)", "1.5", "P4-1", "ingestion/worker.py, ingestion/jobs.py", "미착수", [4]),
    ("P4-3", "GitHub 코드 수집기 구현\n(저장소 복제, 파일 탐색, 지원 확장자 필터링)", "2", "P4-2", "ingestion/loaders/github_loader.py", "미착수", [4]),
    ("P4-4", "문서 파일 수집기 구현\n(텍스트/마크다운 파일 읽기, 정보 추출)", "2", "P4-2", "ingestion/loaders/document_loader.py", "미착수", [4]),
    ("P4-5", "문서 분할 도구 구현\n(일정 크기로 자르기 + 겹침 처리, 키워드 인덱스 생성)", "1.5", "P4-3, P4-4", "ingestion/chunker.py", "미착수", [4, 5]),
    ("P4-6", "분할 -> 벡터 변환 -> DB 저장 처리\n(임베딩 서비스 연동 + 일괄 저장)", "1", "P4-5, P3-4", "ingestion/jobs.py 내 처리 로직", "미착수", [5]),
    ("P4-7", "문서 수집 통합 테스트\n(소스 등록 -> 작업 생성 -> 수집/저장 -> RAG 답변 검증)", "1", "P4-6", "docker compose --profile worker 검증", "미착수", [5]),
]

# ── 데이터 쓰기 ──
row = 2
for task in TASKS:
    task_id, name, duration, dep, deliverable, status, gantt = task

    if task_id == "phase":
        for col in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = PHASE_FILL
            cell.font = PHASE_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER
        ws.cell(row=row, column=1, value="").alignment = CENTER
        ws.cell(row=row, column=2, value=name).alignment = LEFT_WRAP
        ws.cell(row=row, column=6, value=status).alignment = CENTER
        row += 1
        continue

    is_done = status == "완료"
    fill = DONE_FILL if is_done else TODO_FILL

    values = [task_id, name, duration, dep, deliverable, status]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = CELL_FONT
        cell.border = THIN_BORDER
        cell.fill = fill
        cell.alignment = CENTER if col_idx in (1, 3, 4, 6) else LEFT_WRAP

    # 간트 컬럼 (W1~W5)
    for week in range(1, 6):
        col_idx = 6 + week
        cell = ws.cell(row=row, column=col_idx)
        cell.border = THIN_BORDER
        cell.alignment = CENTER

        if is_done and "d1" in gantt:
            cell.fill = GANTT_DONE_FILL
        elif week in gantt:
            cell.fill = GANTT_FILL

    row += 1

# ── 행 높이 설정 ──
for r in range(1, row):
    ws.row_dimensions[r].height = 32

# ── 필터 + 틀 고정 ──
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{row - 1}"
ws.freeze_panes = "A2"

# ── 저장 ──
output_path = "docs/wbs_v2.xlsx"
wb.save(output_path)
print(f"WBS saved: {output_path}")
